import logging
import os

from datavault.settings import ASB_CONFIG, IMPORT_CONFIG, LOGGER
from django.contrib import admin, messages
from django.contrib.admin import AdminSite
from django.contrib.admin.models import ADDITION, CHANGE, LogEntry
from django.contrib.auth.admin import GroupAdmin, UserAdmin
from django.contrib.contenttypes.models import ContentType
from django.http import (FileResponse, HttpResponse, HttpResponseNotFound,
                         HttpResponseRedirect)
from django.utils.translation import gettext_lazy
from gebruikers.models import GebruikersRol
from openpyxl import load_workbook

# from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
# from reportlab.lib.units import inch
# from reportlab.pdfgen import canvas
# from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
# from reportlab.platypus.tableofcontents import TableOfContents
# from reportlab.rl_config import defaultPageSize

from glossary.models import Context, Domein, Gebruiker, GebruikersGroep, SubDomein, Term, TermType

# from .PDF_rapport import RapportTemplate

DEFAULT_PASSWORD = "amsterdamumc"
MAIL_DOMAIN = "@amsterdamumc.nl"

# initialise reportlab (for PDF generation)
# PAGE_HEIGHT = defaultPageSize[1]
# PAGE_WIDTH = defaultPageSize[0]
# styles = getSampleStyleSheet()

# pil_logger = logging.getLogger('PIL')
# pil_logger.setLevel(logging.INFO)

from django.core.exceptions import ObjectDoesNotExist


def email(contact):
    if contact is None or len(contact) == 0: 
        return "dummy" + MAIL_DOMAIN
    details = contact.split() # standaard gescheiden door spaties
    return details[0][0] + "." + details[-1] + MAIL_DOMAIN
 
def updateUser(contact, rol):
    username = contact.strip().replace(" ","") .lower() # geen spaties in een username! 
    try:
        user = Gebruiker.manager.get(username=username)
    except ObjectDoesNotExist:
        user = Gebruiker.manager.create_user(username,email(contact).lower(),DEFAULT_PASSWORD)
        user.first_name = contact.split()[0]
        user.last_name  = " ".join(contact.split()[1:]) 
        user.is_staff   = True
        if rol is not None:  
            user.groups.add(rol)
        user.save()     

def actionImporteerBegrippenlijst(self, request, queryset): 
    filename = os.path.join(ASB_CONFIG["DATA_DIR"],IMPORT_CONFIG["begrippenlijst"]) 
    LOGGER.debug("import gestart vanuit: %s", filename)
    wb = load_workbook(filename)
    # sheetname = wb.sheetnames[1] # de tweede tab
    sheet = wb.worksheets[0]
    # ophalen van default rollen... 
    defaultrol, cr = GebruikersGroep.manager.get_or_create(name__iexact = "Contactpersoon")
    assert(cr is False)
    registrator,cr = GebruikersGroep.manager.get_or_create(name__iexact = "Beheerder begrippenlijst")
    for row in range(2, sheet.max_row):
        rol = None
        naam = sheet.cell(row=row,column=1).value
        if naam is None or len(naam) == 0: 
            continue
        termid = sheet.cell(row=row, column=2).value
        contextnaam = sheet.cell(row=row, column=3).value
        typeterm = sheet.cell(row=row, column=4).value
        typeterm = typeterm.replace("{}","").strip() 
        eenheid = sheet.cell(row=row, column=5).value
        synoniem = sheet.cell(row=row, column=6).value 
        definitie = sheet.cell(row=row, column=7).value
        toelichting = sheet.cell(row=row, column=8).value
        domeinnaam = sheet.cell(row=row, column=11).value
        domeinnaam = domeinnaam.replace("{}","").strip()   
        subdomeinnaam = sheet.cell(row=row, column=12).value
        gebruikersrol = sheet.cell(row=row, column=14).value
        # deze moet er altijd zijn, hij is met de hand toegevoegd... 
        if gebruikersrol is not None: 
            gebruikersrol = gebruikersrol.strip() 
            if len(gebruikersrol) > 0: 
                rol, _  = GebruikersGroep.manager.get_or_create(name__iexact = gebruikersrol)
        beheerdersInfo = sheet.cell(row=row, column=15).value
        if beheerdersInfo is not None: 
            beheerdersInfo = beheerdersInfo.strip() 
            beheerders = beheerdersInfo.split(',')
            for beheerder in beheerders: 
                beheerder = beheerder.strip() 
                if len(beheerder) > 0: 
                    updateUser(beheerder,registrator)

        contactInfo = sheet.cell(row=row, column=13).value
        contacten = contactInfo.split(',') # in de kolom kunnen meerdere gebruikers staan... 
        for contact in contacten: 
            contact = contact.strip() 
            if len(contact) > 0: 
                updateUser(contact, defaultrol)

        context, _ = Context.manager.get_or_create(naam__iexact = contextnaam)
        termType, _ = TermType.manager.get_or_create(naam = typeterm)
        if domeinnaam is not None and len(domeinnaam) > 0:        
            domein, _ = Domein.manager.get_or_create(naam = domeinnaam)
            if subdomeinnaam is not None and len(subdomeinnaam) > 0: 
                subdomein, _ = SubDomein.manager.get_or_create(naam = subdomeinnaam, domein = domein)
                subdomein.domein = domein
                subdomein.save()
        term, _ = Term.manager.get_or_create(id = termid)
        term.id = termid
        term.naam = naam
        term.eenheid = eenheid
        term.synoniem = synoniem
        term.definitie = definitie
        term.toelichting = toelichting
        # koppel de foreign keys
        term.termType = termType
        term.context = context
        term.domein = domein
        term.subdomein = subdomein
        term.rol = rol 
        term.save() 
    LOGGER.debug("lijst met contexts geimporteerd")
    return HttpResponseRedirect(request.META["HTTP_REFERER"])     

actionImporteerBegrippenlijst.short_description = "importeer Begrippenlijst"

class GlossarySite(AdminSite): 
    site_header = gettext_lazy("Dienst ICT - Architectuur, Security en Beleid")

#de admin-site voor het project 
glossary_admin = GlossarySite(name="Woordenboek")

glossary_admin.add_action(actionImporteerBegrippenlijst)

@admin.register(Domein, site=glossary_admin)
class DomeinAdmin(admin.ModelAdmin): 
    list_display = ("naam", )

@admin.register(Term, site=glossary_admin)
class TermAdmin(admin.ModelAdmin):

    list_display = ("naam", "context", "domein", )
    list_filter = ("context", "domein", "rol", "contactpersonen", ) 
    search_fields = ("definitie", "toelichting",) 

    def actionGenereerPDF(self, request, query_set): 
        # file = os.path.join(ASB_CONFIG["REPORT_DIR"],"Begrippenlijst.PDF") 
        # doc = RapportTemplate(file)
        # doc.termSectie(query_set)
        # doc.multiBuild(doc.story)
        # messages.add_message(
        #     request, messages.INFO, "PDF {} gegenereerd".format(file)
        # )
        # LOGGER.debug("PDF {} gegenereerd".format(file))
        return HttpResponseRedirect(request.META["HTTP_REFERER"])        
        
    actions = [actionGenereerPDF]
    actionGenereerPDF.short_description = "Genereer PDF van geselecteerde Termen"

@admin.register(Context, site=glossary_admin)
class ContextAdmin(admin.ModelAdmin): 
    list_display = ('naam',)

@admin.register(LogEntry, site=glossary_admin)
class LogEntryAdmin(admin.ModelAdmin):
    # to have a date-based drilldown navigation in the admin page
    date_hierarchy = 'action_time'

    # to filter the resultes by users, content types and action flags
    list_filter = [
        'user',
        'content_type',
        'action_flag'
    ]

    # when searching the user will be able to search in both object_repr and change_message
    search_fields = [
        'object_repr',
        'change_message'
    ]

    list_display = [
        'action_time',
        'user',
        'content_type',
        'action_flag',
    ]

glossary_admin.register(GebruikersGroep, GroupAdmin) 
glossary_admin.register(Gebruiker, UserAdmin) 
